# Sending Payment Declined (1st, 2nd, Suspended)

import smtplib
import openpyxl as xl
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart

def SignatureType():
	if types[i] == "GPS SmartSole" or types[i] == "Invisabelt":
		Copy = "smartsole@gtxcorp.com"
		team = "SmartSole Team"
		webpage = "https://gpssmartsole.com/"
		return Copy, team, webpage

	elif types[i] == "Track My WorkForce":
		Copy = "tmwf@gtxcorp.com"
		team = "TMWF Team"
		webpage = "https://gtxcorp.com/tmwf-app/"
		return Copy, team, webpage

	else:
		Copy = "take-along@gtxcorp.com"
		team = "Take-Along Tracker Team"
		webpage = "https://gtxcorp.com/take-along-tracker-3g/"
		return Copy, team, webpage

def SubjectType():
	if notices[i] == "1st" or notices[i] == "2nd":
		return f"{OrderIDs[i]} - Payment Declined {notices[i]} notice - {types[i]} Subscription"
	else:
		return f"{OrderIDs[i]} - Service Suspended - {types[i]} Subscription"		

def BodyText():
	if notices[i] == "1st" or notices[i] == "2nd":
		if notices[i] == "1st":
			days = "14"
		else:
			days = "7"

		Body = f"""
<p>Dear {names[i]},</p>
<p>Your recent payment for your&nbsp;{types[i]} Monitoring Service subscription&nbsp;was declined.&nbsp;This could be due to card expiration, change of address, low funds or other card related error.&nbsp;</p>
<p>An active Monitoring&nbsp;subscription is required to continue tracking your device(s).&nbsp;As a courtesy, we have NOT interrupted your Monitoring Service at this time.&nbsp;</p>
<p>This is your <strong>{notices[i]}&nbsp;notice</strong> regarding this issue.&nbsp;To avoid a service interruption please click the link below to repair your account with an updated credit card. You must do this <strong>within&nbsp;{days} business days to avoid the suspension</strong> of your&nbsp;{types[i]} subscription.</p>
<p>Click here to repair your account:</p>
<p><a href="{links[i]}">{links[i]}</a></p>
<p>&nbsp;</p>
<p>Sincerely,</p>
<h3><strong>{team}</strong></h3>
<p><span style="color: #ff0000;"><strong>GTX Corp</strong><br /></span>117 West 9th Street, Suite 1214<strong><br /></strong>Los Angeles, California 90015&nbsp; &nbsp; &nbsp; &nbsp; &nbsp; &nbsp; &nbsp; &nbsp; &nbsp; &nbsp; &nbsp; &nbsp; &nbsp; &nbsp;&nbsp;<br /><a href="mailto:{Copy}">{Copy}<br /></a><a href="{webpage}">{webpage}</a><a href="https://gpssmartsole.com/"><br /></a>www.GTXCorp.com&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;</p>
<p>(OTCQB:&nbsp;GTXO)<br /><span style="color: #0000ff;"><strong><a href="https://www.youtube.com/watch?v=R3Ssy1mnDcA">https://www.youtube.com/watch?v=R3Ssy1mnDcA</a><a href="https://www.youtube.com/watch?v=R3Ssy1mnDcA"><br /></a></strong></span>&nbsp;This message contains confidential information and is intended only for {names[i]}. If you are not the named addressee you should not disseminate, distribute or copy this e-mail. Please notify the sender immediately if you have received this e-mail by mistake and delete this e-mail from your system. Finally, the recipient should check this email and any attachments for the presence of viruses. The company accepts no liability for any damage caused by any virus transmitted by this email.</p>

"""
		return Body
	else:
		Body = f"""
<p>Dear {names[i]},</p>
<p>Your recent payment for your&nbsp;{types[i]} Monitoring Service subscription&nbsp;was declined.&nbsp;This could be due to card expiration, change of address, low funds or other card related error.&nbsp;<br /><br />After several attempts to reach you regarding this issue, <strong>your {types[i]}&nbsp;account has been suspended</strong>. You can reactivate your account by&nbsp;renewing your&nbsp;{types[i]} Monitoring Service subscription&nbsp;with the link below.</p>
<p><strong>Renew your {types[i]}&reg; Monitoring Service Subscription:&nbsp;<br /></strong><a href="{links[i]}">{links[i]}</a></p>
<p class="xxp3" style="margin: 0in; margin-bottom: .0001pt; background: white;"><span style="color: black;">Once renewed, please allow 1 business day for your account to be reactivated.</span></p>
<p class="xxp3" style="margin: 0in; margin-bottom: .0001pt; background: white; caret-color: #000000; font-variant-caps: normal; orphans: auto; text-align: start; widows: auto; -webkit-text-size-adjust: auto; -webkit-text-stroke-width: 0px; background-position: initial initial; background-repeat: initial initial; word-spacing: 0px;"><span style="color: black;">If you have any questions, please contact us at:&nbsp;<a href="mailto:{Copy}">{Copy}</a><span class="apple-converted-space">.</span></span></p>
<p class="xxp3" style="margin: 0in; margin-bottom: .0001pt; background: white; caret-color: #000000; font-variant-caps: normal; orphans: auto; text-align: start; widows: auto; -webkit-text-size-adjust: auto; -webkit-text-stroke-width: 0px; background-position: initial initial; background-repeat: initial initial; word-spacing: 0px;">&nbsp;</p>
<p>Sincerely,</p>
<h3><strong>{team}</strong></h3>
<p><span style="color: #ff0000;"><strong>GTX Corp</strong><br /></span>117 West 9th Street, Suite 1214<strong><br /></strong>Los Angeles, California 90015&nbsp; &nbsp; &nbsp; &nbsp; &nbsp; &nbsp; &nbsp; &nbsp; &nbsp; &nbsp; &nbsp; &nbsp; &nbsp; &nbsp;&nbsp;<br /><a href="mailto:{Copy}">{Copy}<br /></a><a href="{webpage}">{webpage}</a><a href="https://gpssmartsole.com/"><br /></a>www.GTXCorp.com&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;</p>
<p>(OTCQB:&nbsp;GTXO)<br /><span style="color: #0000ff;"><strong><a href="https://www.youtube.com/watch?v=R3Ssy1mnDcA">https://www.youtube.com/watch?v=R3Ssy1mnDcA</a><a href="https://www.youtube.com/watch?v=R3Ssy1mnDcA"><br /></a></strong></span>&nbsp;This message contains confidential information and is intended only for {emails[i]}. If you are not the named addressee you should not disseminate, distribute or copy this e-mail. Please notify the sender immediately if you have received this e-mail by mistake and delete this e-mail from your system. Finally, the recipient should check this email and any attachments for the presence of viruses. The company accepts no liability for any damage caused by any virus transmitted by this email.</p>
<p>&nbsp;</p>

"""
		return Body




###############################################################################################
#>>>>>>>>>>>> INPUT YOUR CREDENTIALS BELOW <<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<
username = "mtulabot@gtxcorp.com"
password = "Mrt_012048554"

###############################################################################################
#>>>>>>>>>>>> LOCATE THE AUTOMATE FOLDER <<<<<<<<<<<<<<<<<<<<<<<<
# example
# wb = xl.load_workbook(r'C:\Users\marzi\Desktop\Automate\PaymentNotice.xlsx')

wb = xl.load_workbook(r'C:\Users\marzi\Desktop\Automate\GTXCorp.xlsx')

print("\nLoading PaymentNotice Workbook...")
payment_notice = wb["PaymentNotice"]
types = []
OrderIDs = []
names = []
emails = []
notices = []
links = []

for col_cells in payment_notice.iter_cols(min_col=1, max_col=1, min_row=2):
	for cell in col_cells: 
		types.append(cell.value)

for col_cells in payment_notice.iter_cols(min_col=2, max_col=2, min_row=2):
	for cell in col_cells: 
		OrderIDs.append(cell.value)

for col_cells in payment_notice.iter_cols(min_col=3, max_col=3, min_row=2):
	for cell in col_cells:
		names.append(cell.value)

for col_cells in payment_notice.iter_cols(min_col=4, max_col=4, min_row=2):
	for cell in col_cells:
		emails.append(cell.value)

for col_cells in payment_notice.iter_cols(min_col=5, max_col=5, min_row=2):
	for cell in col_cells:
		notices.append(cell.value)

for col_cells in payment_notice.iter_cols(min_col=6, max_col=6, min_row=2):
	for cell in col_cells:
		links.append(cell.value)

print("Accessing Workbook Complete!\n\nLogging in to SmartSole Email...\n")

server = smtplib.SMTP("smtp.office365.com", "587")
server.starttls()
server.login(username,password)

for i in range(len(emails)):
	Copy, team, webpage = SignatureType()
	msg = MIMEMultipart()
	msg["From"] = Copy
	msg["To"] = emails[i]
	msg["Cc"] = Copy
	Toaddress = [emails[i], Copy]
	msg["Subject"] = SubjectType()
	Body = BodyText()
	print("Sending messages.")
	# print(emails[i],": ",i+1, " of ",len(emails))
	# print(emails)
	msg.attach(MIMEText(Body, "html")) # HTML is what we want so it could have color and format features
	message = msg.as_string()
	try:
		server.sendmail(username, Toaddress, message) # (sender, receiver, message)
		print(f"Payment Notice sent to {emails[i]}!")
		print(i+1," of ",len(emails),"\n")
	except:
		print("Error has occured!")
		print("Quiting server..")
		break	

server.quit()
print("Quiting the Server...")
print(f"All emails sent successfully!\n\t\tTotal emails sent: {len(emails)}\n")
